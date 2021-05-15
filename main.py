# _*_ coding:utf-8 _*_
from kivy.app import App
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.gridlayout import GridLayout
from kivy.uix.textinput import TextInput
from kivy.config import exists
from random import randint
from time import strftime
import openpyxl as xl
from kivy.core.window import Window
import re


todayChooseMenu = []  # 每次抽取菜单后保存一次
todayChooseDetial = []
schoolText = ''
businessText = ''


class MenuListLayout(GridLayout):  # 第一页中显示菜名和菜单详情的布局（子布局）
    def __init__(self):
        GridLayout.__init__(self, cols=2, spacing=20, padding=20)
        self.menuLabel = Label(text='菜名', font_name='./simhei.ttf')
        self.add_widget(self.menuLabel)
        self.detialLabel = Label(text='详细', font_name='./simhei.ttf')
        self.add_widget(self.detialLabel)
        self.menu1Name = Label(text='', font_name='./simhei.ttf')  # 初始化布局，text全设为'',待菜单抽取出来后由主页面来控制文本改变
        self.add_widget(self.menu1Name)
        self.menu1Detial = Label(text='', font_name='./simhei.ttf')
        self.add_widget(self.menu1Detial)
        self.menu2Name = Label(text='', font_name='./simhei.ttf')
        self.add_widget(self.menu2Name)
        self.menu2Detial = Label(text='', font_name='./simhei.ttf')
        self.add_widget(self.menu2Detial)
        self.menu3Name = Label(text='', font_name='./simhei.ttf')
        self.add_widget(self.menu3Name)
        self.menu3Detial = Label(text='', font_name='./simhei.ttf')
        self.add_widget(self.menu3Detial)
        self.menu4Name = Label(text='', font_name='./simhei.ttf')
        self.add_widget(self.menu4Name)
        self.menu4Detial = Label(text='', font_name='./simhei.ttf')
        self.add_widget(self.menu4Detial)
        self.menu5Name = Label(text='', font_name='./simhei.ttf')
        self.add_widget(self.menu5Name)
        self.menu5Detial = Label(text='', font_name='./simhei.ttf')
        self.add_widget(self.menu5Detial)


class BusinessLayout(GridLayout):  # 给商家的菜单的布局
    def __init__(self):
        width, height = Window.size
        GridLayout.__init__(self, cols=1, spacing=20, padding=20)
        self.topLabel = Label(text='菜单', font_name='./simhei.ttf', height=180*height/3168, size_hint_y=None)
        self.add_widget(self.topLabel)
        self.mainLabel = Label(text='', font_name='./simhei.ttf')
        self.getText()
        self.add_widget(self.mainLabel)
        self.backButton = Button(text='返回', font_name='./simhei.ttf', height=240*height/3168, size_hint_y=None)
        self.add_widget(self.backButton)
        self.backButton.bind(on_press=self.backEvent)

    def getText(self):
        global todayChooseDetial
        numList = {}
        if len(todayChooseDetial) != 0:
            tempText = ''
            for sts in todayChooseDetial:
                # ————————提取菜名的部分————————————
                sts = sts.replace(' ', '')
                if len(re.findall('（', sts)) != 0:
                    sts = sts.replace('（', '(')
                if len(re.findall('）', sts)) != 0:
                    sts = sts.replace('）', ')')
                ls = re.findall(r'[)]*(\D+)[(]\d+[g克斤个袋]+[)]', sts)
                for index in range(len(ls)):
                    if len(re.findall(r'[)]', ls[index])) != 0:
                        ls[index] = re.findall(r'[)](\D+)', ls[index])[0]
                # ————————提取菜名的部分————————————

                # ————————提取数量的部分————————————
                num = re.findall(r'(\d+[斤个袋]+)', sts)
                # ————————提取数量的部分————————————

                for index in range(len(ls)):
                    flag = ls[index]
                    if flag not in numList.keys():
                        numList[flag] = num[index]
                    else:
                        numList[flag] = str(eval(numList[flag][:-1]) + eval(num[index][:-1]))+numList[flag][-1]
            for flag in numList.keys():
                tempText += flag+': '+numList[flag]+'\n'
            self.mainLabel.text = tempText

    def backEvent(self, event):  # 返回第一页
        self.parent.add_widget(LogicScreen())
        self.parent.remove_widget(self)


class SchoolLayout(GridLayout):  # 给学校的菜单的布局
    def __init__(self):
        width, height = Window.size
        GridLayout.__init__(self, cols=1, spacing=20, padding=20)
        self.topLabel = Label(text='菜单', font_name='./simhei.ttf', height=180*height/3168, size_hint_y=None)
        self.add_widget(self.topLabel)
        self.mainLabel = Label(text='', font_name='./simhei.ttf')
        self.getText()
        self.add_widget(self.mainLabel)
        self.backButton = Button(text='返回', font_name='./simhei.ttf', height=240*height/3168, size_hint_y=None)
        self.add_widget(self.backButton)
        self.backButton.bind(on_press=self.backEvent)

    def getText(self):
        global todayChooseDetial
        numList = {}
        if len(todayChooseDetial) != 0:
            tempText = ''
            numList = {}
            for sts in todayChooseDetial:
                # ————————提取菜名的部分————————————
                sts = sts.replace(' ', '')
                if len(re.findall('（', sts)) != 0:
                    sts = sts.replace('（', '(')
                if len(re.findall('）', sts)) != 0:
                    sts = sts.replace('）', ')')
                ls = re.findall(r'[)]*(\D+)[(]\d+[.]*\d*[千k]*[g克斤个袋]*[)]', sts)
                for index in range(len(ls)):
                    if len(re.findall(r'[)]', ls[index])) != 0:
                        ls[index] = re.findall(r'[)](\D+)', ls[index])[0]
                # ————————提取菜名的部分————————————

                # ————————提取数量的部分————————————
                num = re.findall(r'(\d+[.]*\d*.?[克g]+)', sts)
                for flag in range(len(num)):
                    st = num[flag]
                    neo = ''
                    if len(re.findall(r'千克', st)) != 0 or len(re.findall(r'kg', st)) != 0:
                        number = eval(st[:-2]) * 1000
                        if number % 1 == 0:
                            neo = str(int(number)) + '克'
                        else:
                            neo = str(number) + '克'
                        num[flag] = neo
                del st, neo
                # ————————提取数量的部分————————————

                for index in range(len(ls)):
                    flag = ls[index]
                    if flag not in numList.keys():
                        numList[flag] = num[index]
                    else:
                        numList[flag] = str(eval(numList[flag][:-1]) + eval(num[index][:-1])) + numList[flag][-1]
            for flag in numList.keys():
                tempText += flag+': '+numList[flag]+'\n'
            self.mainLabel.text = tempText

    def backEvent(self, event):  # 返回第一页
        self.parent.add_widget(LogicScreen())
        self.parent.remove_widget(self)


class LogicScreen(GridLayout):  # 第一页的主布局，
    def __init__(self):
        GridLayout.__init__(self, cols=1, spacing=20, padding=20)
        width, height = Window.size
        self.topLabel = Label(text='菜单', font_name='./simhei.ttf', height=200*height/3168, size_hint_y=None)  # 菜单
        self.add_widget(self.topLabel)
        self.dicButton = Button(text='确认', font_name='./simhei.ttf', height=180*height/3168, size_hint_y=None,
                                size_hint_x=0.8)
        self.add_widget(self.dicButton)
        self.dicButton.bind(on_press=self.dicEvent)
        self.menuTable = MenuListLayout()
        self.add_widget(self.menuTable)
        self.initialMenu()
        self.button1 = Button(text='抽取菜单', font_name='./simhei.ttf', height=180*height/3168, size_hint_y=None)
        self.add_widget(self.button1)
        self.button1.bind(on_press=self.callBack)
        self.newPageButton1 = Button(text='商家用', font_name='./simhei.ttf', height=200*height/3168, size_hint_y=None)
        self.add_widget(self.newPageButton1)
        self.newPageButton2 = Button(text='学校用', font_name='./simhei.ttf', height=200*height/3168, size_hint_y=None)
        self.add_widget(self.newPageButton2)
        self.newPageButton1.bind(on_press=self.toBusiness)  # 给商家的
        self.newPageButton2.bind(on_press=self.toSchool)  # 给学校的
        self.nextPageButton = Button(text='下一页', font_name='./simhei.ttf', height=240*height/3168, size_hint_y=None)
        self.add_widget(self.nextPageButton)
        self.nextPageButton.bind(on_press=self.toNextPage)  # 菜单输入界面
        self.length = 8
        self.menuD = {}  # 菜名的字典
        self.detailedD = {}  # 详细的字典
        self.checkMenu()

    def initialMenu(self):  # 重返第一页的时候重新在文本上打印刚才抽取的文本
        global todayChooseMenu, todayChooseDetial
        if len(todayChooseDetial) != 0:
            self.menuTable.menu1Name.text = todayChooseMenu[0]
            self.menuTable.menu2Name.text = todayChooseMenu[1]
            self.menuTable.menu3Name.text = todayChooseMenu[2]
            self.menuTable.menu4Name.text = todayChooseMenu[3]
            self.menuTable.menu5Name.text = todayChooseMenu[4]
            tlk = todayChooseDetial.copy()
            for index in range(len(tlk)):
                sts = tlk[index]
                sts = sts.replace(' ', '')
                count = 0
                neo = ''
                for flag in range(len(sts)):
                    neo += sts[flag]
                    if sts[flag] in [')', '）']:
                        count += 1
                    if count == 2:
                        count = 0
                        neo += '\n'
                tlk[index] = neo
            self.menuTable.menu1Detial.text = tlk[0]
            self.menuTable.menu2Detial.text = tlk[1]
            self.menuTable.menu3Detial.text = tlk[2]
            self.menuTable.menu4Detial.text = tlk[3]
            self.menuTable.menu5Detial.text = tlk[4]

    def toBusiness(self, event):  # 跳转到给商家的菜单
        self.parent.add_widget(BusinessLayout())
        self.parent.remove_widget(self)

    def toSchool(self, event):
        self.parent.add_widget(SchoolLayout())
        self.parent.remove_widget(self)

    def checkMenu(self):  # 打开app的时候读取已有菜单
        if not exists('./item.xlsx'):
            tempWorkBook = xl.Workbook()  # 添加try以防缺失初始菜单
            read_sheet = tempWorkBook.create_sheet('菜单')
            if 'Sheet' in tempWorkBook.get_sheet_names():
                del tempWorkBook['Sheet']
            menuDic = {0: '西红柿炒蛋',
                       1: '番茄炒蛋',
                       2: '蛋炒番茄',
                       3: '蛋炒西红柿',
                       4: '炒蛋',
                       5: '凉拌番茄',
                       6: '蛋炒饭',
                       7: '饭炒蛋'}
            detialDic = {0: '鸡蛋(50克)(5个) 西红柿(130克)(3个)',
                         1: '鸡蛋(50克)(5个) 西红柿(130克)(3个)',
                         2: '鸡蛋(50克)(5个) 西红柿(130克)(3个)',
                         3: '鸡蛋(50克)(5个) 西红柿(130克)(3个)',
                         4: '鸡蛋(50克)(5个) ',
                         5: '西红柿(130克)(3个)',
                         6: '鸡蛋(60克)(6个) 大米(120克)(60斤)',
                         7: '鸡蛋(60克)(6个) 大米(120克)(60斤)'}
            read_sheet.cell(row=1, column=1, value='菜名')
            read_sheet.cell(row=1, column=2, value='详细')
            for i in range(8):
                read_sheet.cell(row=i + 2, column=1, value=menuDic[i])
                read_sheet.cell(row=i + 2, column=2, value=detialDic[i])
            tempWorkBook.save('./item.xlsx')
            del tempWorkBook, read_sheet, menuDic, detialDic
        readWorkBook = xl.load_workbook('./item.xlsx')
        sheet = readWorkBook['菜单']
        maxRow = sheet.max_row
        for index in range(maxRow-1):
            self.menuD[index] = sheet.cell(row=index+2, column=1).value  # 循环获取菜名
            self.detailedD[index] = sheet.cell(row=index + 2, column=2).value  # 循环获取详情
        self.length = maxRow-1

    def toNextPage(self, event):  # 第一页到输入新菜单的界面
        self.parent.add_widget(SecondScreen())
        self.parent.remove_widget(self)

    def callBack(self, event):  # 随机抽取菜单  这是抽取菜单的按钮
        global todayChooseMenu, todayChooseDetial
        lis = list(range(self.length))
        temp1 = lis.pop(randint(0, len(lis)-1))
        temp2 = lis.pop(randint(0, len(lis)-1))
        temp3 = lis.pop(randint(0, len(lis)-1))
        temp4 = lis.pop(randint(0, len(lis)-1))
        temp5 = lis.pop(randint(0, len(lis)-1))
        tlm = [self.menuD[temp1], self.menuD[temp2], self.menuD[temp3], self.menuD[temp4], self.menuD[temp5]]
        tlk = [self.detailedD[temp1], self.detailedD[temp2], self.detailedD[temp3], self.detailedD[temp4],
               self.detailedD[temp5]]
        todayChooseMenu = tlm.copy()
        todayChooseDetial = tlk.copy()
        for index in range(len(tlk)):
            sts = tlk[index]
            sts = sts.replace(' ', '')
            count = 0
            neo = ''
            for flag in range(len(sts)):
                neo += sts[flag]
                if sts[flag] in [')', '）']:
                    count += 1
                if count == 2:
                    count = 0
                    neo += '\n'
            tlk[index] = neo
        self.menuTable.menu1Name.text, self.menuTable.menu1Detial.text = self.menuD[temp1], tlk[0]
        self.menuTable.menu2Name.text, self.menuTable.menu2Detial.text = self.menuD[temp2], tlk[1]
        self.menuTable.menu3Name.text, self.menuTable.menu3Detial.text = self.menuD[temp3], tlk[2]
        self.menuTable.menu4Name.text, self.menuTable.menu4Detial.text = self.menuD[temp4], tlk[3]
        self.menuTable.menu5Name.text, self.menuTable.menu5Detial.text = self.menuD[temp5], tlk[4]
        return self.button1

    def dicEvent(self, event):  # 确认按钮的控件
        self.saveData()
        return self.dicButton

    def saveData(self):  # 确认菜单后通过该方法保存菜单至menu.xlsx
        global todayChooseMenu, todayChooseDetial
        if len(todayChooseDetial) != 0 and len(todayChooseMenu) != 0:
            # ——————————已抽取菜单里新增今日的菜单——————————————————
            try:
                workbook = xl.load_workbook('./menu.xlsx')  # /storage/emulated/0/Documents/menu.xlsx
            except FileNotFoundError:
                workbook = xl.Workbook()
            names = workbook.get_sheet_names()

            if 'Sheet' in names:
                del workbook['Sheet']

            if strftime("%Y") + '年的菜单' not in names:
                sheet = workbook.create_sheet(strftime("%Y") + '年的菜单')
            else:
                sheet = workbook[strftime("%Y") + '年的菜单']

            maxRow = sheet.max_row
            sheet.cell(row=maxRow + 1, column=1, value=strftime('%m') + '月' + strftime('%d') + '日')
            for index in range(5):
                sheet.cell(row=maxRow + index + 2, column=1, value=todayChooseMenu[index])
                sheet.cell(row=maxRow + index + 2, column=2, value=todayChooseDetial[index])

            workbook.save('./menu.xlsx')
            # ——————————已抽取菜单里新增今日的菜单——————————————————

            # ————————————创建今天的菜单——————————————————
            workbook = xl.Workbook()

            del workbook['Sheet']
            sheet = workbook.create_sheet('Sheet')

            sheet.cell(row=1, column=1, value=strftime('%m') + '月' + strftime('%d') + '日')
            for index in range(5):
                sheet.cell(row=index + 2, column=1, value=todayChooseMenu[index])
                sheet.cell(row=index + 2, column=2, value=todayChooseDetial[index])

            del sheet

            # ————————————保存给商家的部分——————————————
            sheet = workbook.create_sheet('商家')

            numList = {}
            if len(todayChooseDetial) != 0:
                tempText = ''
                for sts in todayChooseDetial:
                    # ————————提取菜名的部分————————————
                    sts = sts.replace(' ', '')
                    if len(re.findall('（', sts)) != 0:
                        sts = sts.replace('（', '(')
                    if len(re.findall('）', sts)) != 0:
                        sts = sts.replace('）', ')')
                    ls = re.findall(r'[)]*(\D+)[(]\d+[g克斤个袋]+[)]', sts)
                    for index in range(len(ls)):
                        if len(re.findall(r'[)]', ls[index])) != 0:
                            ls[index] = re.findall(r'[)](\D+)', ls[index])[0]
                    # ————————提取菜名的部分————————————

                    # ————————提取数量的部分————————————
                    num = re.findall(r'(\d+[斤个袋]+)', sts)
                    # ————————提取数量的部分————————————

                    for index in range(len(ls)):
                        flag = ls[index]
                        if flag not in numList.keys():
                            numList[flag] = num[index]
                        else:
                            numList[flag] = str(eval(numList[flag][:-1]) + eval(num[index][:-1])) + numList[flag][-1]
                sheet.cell(row=1, column=1, value=strftime('%m') + '月' + strftime('%d') + '日')
                index = 2
                for flag in numList.keys():
                    tempText = flag + ': ' + numList[flag]
                    sheet.cell(row=index, column=1, value=tempText)
                    index += 1

            del sheet
            # ————————————保存给商家的部分————————————————

            # ————————————保存给学校的部分————————————————
            sheet = workbook.create_sheet('学校')

            numList = {}
            if len(todayChooseDetial) != 0:
                tempText = ''
                numList = {}
                for sts in todayChooseDetial:
                    # ————————提取菜名的部分————————————
                    sts = sts.replace(' ', '')
                    if len(re.findall('（', sts)) != 0:
                        sts = sts.replace('（', '(')
                    if len(re.findall('）', sts)) != 0:
                        sts = sts.replace('）', ')')
                    ls = re.findall(r'[)]*(\D+)[(]\d+[.]*\d*[千k]*[g克斤个袋]*[)]', sts)
                    for index in range(len(ls)):
                        if len(re.findall(r'[)]', ls[index])) != 0:
                            ls[index] = re.findall(r'[)](\D+)', ls[index])[0]
                    # ————————提取菜名的部分————————————

                    # ————————提取数量的部分————————————
                    num = re.findall(r'(\d+[.]*\d*.?[克g]+)', sts)
                    for flag in range(len(num)):
                        st = num[flag]
                        neo = ''
                        if len(re.findall(r'千克', st)) != 0 or len(re.findall(r'kg', st)) != 0:
                            number = eval(st[:-2]) * 1000
                            if number % 1 == 0:
                                neo = str(int(number)) + '克'
                            else:
                                neo = str(number) + '克'
                            num[flag] = neo
                    del st, neo
                    # ————————提取数量的部分————————————

                    for index in range(len(ls)):
                        flag = ls[index]
                        if flag not in numList.keys():
                            numList[flag] = num[index]
                        else:
                            numList[flag] = str(eval(numList[flag][:-1]) + eval(num[index][:-1])) + numList[flag][-1]
                sheet.cell(row=1, column=1, value=strftime('%m') + '月' + strftime('%d') + '日')
                index = 2
                for flag in numList.keys():
                    tempText = flag + ': ' + numList[flag]
                    sheet.cell(row=index, column=1, value=tempText)
                    index += 1

            del sheet
            # ————————————保存给学校的部分————————————————

            workbook.save('./todayMenu.xlsx')
            # ————————————创建今天的菜单——————————————————

        else:
            self.menuTable.menu1Name.text = '还未抽取菜单'


class InputMenu(GridLayout):  # 第二页的文本输入框
    def __init__(self):
        GridLayout.__init__(self, cols=2, spacing=20, padding=20)
        self.menuInputLabel = Label(text='菜单', font_name='./simhei.ttf')
        self.add_widget(self.menuInputLabel)
        self.detialInputLabel = Label(text='详细信息', font_name='./simhei.ttf')
        self.add_widget(self.detialInputLabel)
        self.menuInput = TextInput(font_name='./simhei.ttf')
        self.add_widget(self.menuInput)
        self.detialInput = TextInput(font_name='./simhei.ttf')
        self.add_widget(self.detialInput)


class SecondScreen(GridLayout):  # 第二页
    def __init__(self):
        GridLayout.__init__(self, cols=1, spacing=20, padding=20)
        width, height = Window.size
        self.topLabel = Label(text='新菜单输入', height=200*height/3168, size_hint_y=None, font_name='./simhei.ttf')
        self.add_widget(self.topLabel)
        self.newMenuInputLayout = InputMenu()
        self.add_widget(self.newMenuInputLayout)
        self.decButton = Button(text='确认', font_name='./simhei.ttf', height=220*height/3168, size_hint_y=None)
        self.add_widget(self.decButton)
        self.decButton.bind(on_press=self.decideButtonEvent)
        self.backButton = Button(text='返回', font_name='./simhei.ttf', height=240*height/3168, size_hint_y=None)
        self.add_widget(self.backButton)
        self.backButton.bind(on_press=self.touchBack)

    def touchBack(self, event):  # 第二页的返回键
        self.parent.add_widget(LogicScreen())
        self.parent.remove_widget(self)

    def decideButtonEvent(self, event):  # 第二页的确认新菜单的输入按键
        self.saveNewMenu()
        return self.decButton

    def saveNewMenu(self):  # 新菜单存入xlsx的方法
        menuText = self.newMenuInputLayout.menuInput.text
        detialText = self.newMenuInputLayout.detialInput.text
        menuText = menuText.replace(' ', '')
        detialText = detialText.replace(' ', '')
        if menuText == '' or detialText == '':
            pass
        else:
            workBook = xl.load_workbook('./item.xlsx')
            sheet = workBook['菜单']
            maxRow = sheet.max_row
            sheet.cell(row=maxRow+1, column=1, value=menuText)
            sheet.cell(row=maxRow+1, column=2, value=detialText)
            workBook.save('./item.xlsx')
            del menuText, detialText, workBook, sheet, maxRow


class MyApp(App):
    def build(self):
        return LogicScreen()


if __name__ == '__main__':
    test = MyApp()
    test.run()
